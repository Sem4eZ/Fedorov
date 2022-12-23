import csv
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
import matplotlib.pyplot as plt
import numpy as np


class Vacancy:
    currency_to_rub = {
        "RUR": 1,
        "BYR": 23.91,
        "KZT": 0.13,
        "EUR": 59.90,
        "KGS": 0.76,
        "GEL": 21.74,
        "UZS": 0.0055,
        "UAH": 1.64,
        "USD": 60.66,
        "AZN": 35.68
    }

    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average_value = self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2


class DataSet:
    def __init__(self, name_file, name_vacancy):
        self.name_file = name_file
        self.name_vacancy = name_vacancy

    @staticmethod
    def addition(sum_value, dic, key):
        if key not in dic:
            dic[key] = sum_value
        else:
            dic[key] += sum_value

    def csv_reader(self):
        with open(self.name_file, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            headings = next(reader)
            for inf in reader:
                if len(headings) == len(inf) and '' not in inf:
                    yield dict(zip(headings, inf))

    @staticmethod
    def make_average_value(dic):
        average_dic = {}
        for key, values in dic.items():
            average_dic[key] = int(sum(values) / len(values))
        return average_dic

    @staticmethod
    def output_statistics(first_statistical_data, second_statistical_data, third_statistical_data,
                          fourth_statistical_data,
                          fifth_statistical_data, sixth_statistical_data):
        print('Динамика уровня зарплат по годам: {0}'.format(first_statistical_data))
        print('Динамика количества вакансий по годам: {0}'.format(second_statistical_data))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(third_statistical_data))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(fourth_statistical_data))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(fifth_statistical_data))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(sixth_statistical_data))

    def make_statistic(self):
        vacancies_count = 0
        wages = {}
        city_wages = {}
        vacancy_wages = {}
        all_vacancies = self.csv_reader()

        for vacancy_dict in all_vacancies:
            vacancy = Vacancy(vacancy_dict)
            self.addition([vacancy.salary_average_value], wages, vacancy.year)
            if vacancy.name.find(self.name_vacancy) != -1:
                self.addition([vacancy.salary_average_value], vacancy_wages, vacancy.year)
            vacancies_count += 1
            self.addition([vacancy.salary_average_value], city_wages, vacancy.area_name)
        list_for_dict_by_name = [(key, len(value)) for key, value in vacancy_wages.items()]
        vacancies_num_by_name = dict(list_for_dict_by_name)
        list_for_dict_num = [(key, len(value)) for key, value in wages.items()]
        vacancies_num = dict(list_for_dict_num)

        if not vacancy_wages:
            list_for_dict_by_name = [(key, 0) for key, value in vacancies_num.items()]
            vacancies_num_by_name = dict(list_for_dict_by_name)
            list_for_dict_wages = [(key, [0]) for key, value in wages.items()]
            vacancy_wages = dict(list_for_dict_wages)
        third_statistical_data = self.make_average_value(city_wages)
        second_statistical_data = self.make_average_value(vacancy_wages)
        statistical_data = self.make_average_value(wages)

        fourth_statistical_data = {}
        for year, year_salaries in city_wages.items():
            fourth_statistical_data[year] = round(len(year_salaries) / vacancies_count, 4)
        fourth_statistical_data = list(
            filter(lambda s: s[-1] >= 0.01, [(key, value) for key, value in fourth_statistical_data.items()]))
        fourth_statistical_data.sort(key=lambda s: s[-1], reverse=True)
        fifth_statistical_data = fourth_statistical_data.copy()
        fourth_statistical_data = dict(fourth_statistical_data)
        third_statistical_data = list(filter(lambda s: s[0] in list(fourth_statistical_data.keys()),
                                             [(key, value) for key, value in third_statistical_data.items()]))
        third_statistical_data.sort(key=lambda s: s[-1], reverse=True)
        third_statistical_data = dict(third_statistical_data[:10])
        fifth_statistical_data = dict(fifth_statistical_data[:10])

        return statistical_data, vacancies_num, second_statistical_data, vacancies_num_by_name, third_statistical_data, fifth_statistical_data


class InputConnect:
    def __init__(self):
        self.name_file = input('Введите название файла: ')
        self.name_vacancy = input('Введите название профессии: ')

        data_set = DataSet(self.name_file, self.name_vacancy)
        first_statistical_data, second_statistical_data, third_statistical_data, fourth_statistical_data, fifth_statistical_data, sixth_statistical_data = data_set.make_statistic()
        data_set.output_statistics(first_statistical_data, second_statistical_data, third_statistical_data,
                                   fourth_statistical_data,
                                   fifth_statistical_data, sixth_statistical_data)

        report = Report(self.name_vacancy, first_statistical_data, second_statistical_data, third_statistical_data,
                        fourth_statistical_data, fifth_statistical_data, sixth_statistical_data)
        report.generate_excel()
        report.generate_image()


class Report:
    def __init__(self, name_vacancy, first_statistical_data, second_statistical_data, third_statistical_data,
                 fourth_statistical_data,
                 fifth_statistical_data, sixth_statistical_data):
        self.workbook = Workbook()
        self.sixth_statistical_data = sixth_statistical_data
        self.fifth_statistical_data = fifth_statistical_data
        self.fourth_statistical_data = fourth_statistical_data
        self.third_statistical_data = third_statistical_data
        self.second_statistical_data = second_statistical_data
        self.first_statistical_data = first_statistical_data
        self.name_vacancy = name_vacancy

    @staticmethod
    def make_widths(informations):
        widths = []
        for inf in informations:
            for i, cell in enumerate(inf):
                cell = str(cell)
                if len(widths) <= i:
                    widths += [len(cell)]
                else:
                    if len(cell) > widths[i]:
                        widths[i] = len(cell)
        return widths

    @staticmethod
    def formatting_tabs(first_tab, second_tab, informations, first_statistical_data, fifth_statistical_data):
        for index, s in enumerate(fifth_statistical_data):
            second_tab['E' + str(index + 2)].number_format = '0.00%'

        for column in 'ABCDE':
            first_tab[column + '1'].font = Font(bold=True)
            second_tab[column + '1'].font = Font(bold=True)

        for row, s in enumerate(first_statistical_data):
            for column in 'ABCDE':
                first_tab[column + str(row + 1)].border = Border(left=Side(border_style='thin', color='00000000'),
                                                                 bottom=Side(border_style='thin', color='00000000'),
                                                                 right=Side(border_style='thin', color='00000000'),
                                                                 top=Side(border_style='thin', color='00000000'))

        for row in range(len(informations)):
            for column in 'ABDE':
                second_tab[column + str(row + 1)].border = Border(left=Side(border_style='thin', color='00000000'),
                                                                  bottom=Side(border_style='thin', color='00000000'),
                                                                  right=Side(border_style='thin', color='00000000'),
                                                                  top=Side(border_style='thin', color='00000000'))

    def generate_excel(self):
        first_tab = self.workbook.active
        first_tab.title = 'Статистика по годам'
        first_tab.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.name_vacancy, 'Количество вакансий',
                          'Количество вакансий - ' + self.name_vacancy])
        for year in self.first_statistical_data.keys():
            first_tab.append([year, self.first_statistical_data[year], self.third_statistical_data[year],
                              self.second_statistical_data[year],
                              self.fourth_statistical_data[year]])

        informations = [
            ['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.name_vacancy, ' Количество вакансий',
             ' Количество вакансий - ' + self.name_vacancy]]

        widths = self.make_widths(informations)

        for i, width in enumerate(widths, 1):
            first_tab.column_dimensions[get_column_letter(i)].width = width + 2

        informations = []
        informations.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (first_city, first_value), (second_city, second_value) in zip(self.fifth_statistical_data.items(),
                                                                          self.sixth_statistical_data.items()):
            informations.append([first_city, first_value, '', second_city, second_value])
        second_tab = self.workbook.create_sheet('Статистика по городам')

        for inf in informations:
            second_tab.append(inf)

        widths = self.make_widths(informations)

        for i, width in enumerate(widths, 1):
            second_tab.column_dimensions[get_column_letter(i)].width = width + 2

        self.formatting_tabs(first_tab, second_tab, informations, self.first_statistical_data,
                             self.fifth_statistical_data)
        self.workbook.save('report.xlsx')

    @staticmethod
    def make_axes(ax, first_bar, second_bar, name_vacancy, statistical_data, first_legend_str, second_legend_str):
        ax.grid(axis='y')
        ax.legend((first_bar[0], second_bar[0]), (first_legend_str, second_legend_str + name_vacancy.lower()),
                  prop={'size': 8})

        ax.set_xticks(np.array(list(statistical_data.keys())) - 0.2,
                      list(statistical_data.keys()), rotation=90)
        ax.xaxis.set_tick_params(labelsize=8)
        ax.yaxis.set_tick_params(labelsize=8)

    def generate_image(self):
        fig, ((first_ax, second_ax), (third_ax, fourth_ax)) = plt.subplots(nrows=2, ncols=2)

        first_bar = first_ax.bar(np.array(list(self.first_statistical_data.keys())) - 0.4,
                                 self.first_statistical_data.values(), width=0.4)
        second_bar = first_ax.bar(np.array(list(self.first_statistical_data.keys())),
                                  self.third_statistical_data.values(), width=0.4)
        first_ax.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})

        self.make_axes(first_ax, first_bar, second_bar, self.name_vacancy, self.first_statistical_data, 'средняя з/п',
                       'з/п ')

        second_ax.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        first_bar = second_ax.bar(np.array(list(self.second_statistical_data.keys())) - 0.4,
                                  self.second_statistical_data.values(), width=0.4)
        second_bar = second_ax.bar(np.array(list(self.second_statistical_data.keys())),
                                   self.fourth_statistical_data.values(), width=0.4)

        self.make_axes(second_ax, first_bar, second_bar, self.name_vacancy, self.second_statistical_data,
                       'Количество вакансий', 'Количество вакансий\n')

        third_ax.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        third_ax.barh(list([str(a).replace(' ', '\n').replace('-', '-\n') for a in
                            reversed(list(self.fifth_statistical_data.keys()))]),
                      list(reversed(list(self.fifth_statistical_data.values()))), color='blue', height=0.5,
                      align='center')
        third_ax.grid(axis='x')
        third_ax.xaxis.set_tick_params(labelsize=8)
        third_ax.yaxis.set_tick_params(labelsize=6)

        fourth_ax.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        remainder_list = [val for val in self.sixth_statistical_data.values()]
        remainder = 1 - sum(remainder_list)
        self.sixth_statistical_data = dict(sorted(self.sixth_statistical_data.items(), key=lambda x: x[1]))
        fourth_ax.pie(list(self.sixth_statistical_data.values()) + [remainder],
                      labels=list(self.sixth_statistical_data.keys()) + ['Другие'],
                      textprops={'fontsize': 6})
        plt.tight_layout()
        plt.savefig('graph.png')


InputConnect()
